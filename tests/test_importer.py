"""
Importer tests — csv_import.py compatibiliteit.

Dekt:
  - Crossuite-stijl: aparte VOORNAAM/ACHTERNAAM kolommen
  - MyOrganizer-stijl: volledige naam in één kolom
  - Multi-sheet Excel: kiest sheet met meeste e-mails
  - E-maildetectie via cel-scan
  - Datumdetectie DD/MM/YYYY

Run:
    pytest tests/test_importer.py -v
"""

import csv
import io
import os
import sys
import tempfile
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))
from csv_import import _process_rows, load_csv_file


# ── Helpers ──────────────────────────────────────────────────────────────────

def _make_csv_rows(headers, rows):
    """Build DictReader-style list from headers + list-of-lists."""
    return [dict(zip(headers, r)) for r in rows]


# ── 1. Crossuite-stijl: hoofdletters ─────────────────────────────────────────

class TestCrossuiteSplitNames:
    HEADERS = ['DATUM', 'VOORNAAM', 'ACHTERNAAM', 'E-MAIL', 'ITEMS']

    def _row(self, datum, voornaam, achternaam, email, items='Osteopathie'):
        return [datum, voornaam, achternaam, email, items]

    def test_naam_gecombineerd(self):
        data = _make_csv_rows(self.HEADERS,
                              [self._row('22/05/2026', 'Lisa', 'Van den Abeele',
                                         'test@example.com')])
        rows, _ = _process_rows(self.HEADERS, data, 'test.csv')
        assert rows[0]['naam'] == 'Van den Abeele Lisa'

    def test_voornaam_correct(self):
        data = _make_csv_rows(self.HEADERS,
                              [self._row('22/05/2026', 'Lisa', 'Van den Abeele',
                                         'test@example.com')])
        rows, _ = _process_rows(self.HEADERS, data, 'test.csv')
        assert rows[0]['voornaam'] == 'Lisa'

    def test_achternaam_correct(self):
        data = _make_csv_rows(self.HEADERS,
                              [self._row('22/05/2026', 'Lisa', 'Van den Abeele',
                                         'test@example.com')])
        rows, _ = _process_rows(self.HEADERS, data, 'test.csv')
        assert rows[0]['achternaam'] == 'Van den Abeele'

    def test_email_correct(self):
        data = _make_csv_rows(self.HEADERS,
                              [self._row('22/05/2026', 'Jan', 'Pieters',
                                         'jan.pieters@example.be')])
        rows, _ = _process_rows(self.HEADERS, data, 'test.csv')
        assert rows[0]['email'] == 'jan.pieters@example.be'

    def test_datum_consult_correct(self):
        data = _make_csv_rows(self.HEADERS,
                              [self._row('15/03/2026', 'Anna', 'De Smet',
                                         'anna@example.com')])
        rows, _ = _process_rows(self.HEADERS, data, 'test.csv')
        assert rows[0]['datum_consult'] == '2026-03-15'

    def test_meerdere_rijen(self):
        data = _make_csv_rows(self.HEADERS, [
            self._row('01/05/2026', 'Lisa',  'Van den Abeele', 'a@example.com'),
            self._row('02/05/2026', 'Johan', 'Martens',        'b@example.com'),
        ])
        rows, stats = _process_rows(self.HEADERS, data, 'test.csv')
        assert len(rows) == 2
        assert stats['rijen_ok'] == 2
        assert rows[1]['naam'] == 'Martens Johan'

    def test_rij_zonder_email_overgeslagen(self):
        data = _make_csv_rows(self.HEADERS, [
            self._row('01/05/2026', 'Lisa', 'Van den Abeele', ''),
            self._row('02/05/2026', 'Jan',  'Pieters',        'jan@example.com'),
        ])
        rows, stats = _process_rows(self.HEADERS, data, 'test.csv')
        assert len(rows) == 1
        assert stats['rijen_geen_email'] == 1

    def test_lege_rij_overgeslagen(self):
        data = _make_csv_rows(self.HEADERS, [
            ['', '', '', '', ''],
            self._row('01/05/2026', 'Anna', 'De Smet', 'anna@example.com'),
        ])
        rows, stats = _process_rows(self.HEADERS, data, 'test.csv')
        assert len(rows) == 1
        assert stats['rijen_leeg'] == 1


# ── 2. Crossuite kleine letters / gemengde kolomnamen ────────────────────────

class TestCrossuiteKleineLetters:
    def test_voornaam_kleine_letters(self):
        headers = ['datum', 'Voornaam', 'Achternaam', 'E-mail']
        data = _make_csv_rows(headers,
                              [['22/05/2026', 'Sara', 'Willems', 'sara@example.com']])
        rows, _ = _process_rows(headers, data, 'test.csv')
        assert rows[0]['voornaam'] == 'Sara'
        assert rows[0]['naam'] == 'Willems Sara'

    def test_first_name_last_name_english(self):
        headers = ['date', 'first name', 'last name', 'email']
        data = _make_csv_rows(headers,
                              [['2026-05-22', 'John', 'Smith', 'john@example.com']])
        rows, _ = _process_rows(headers, data, 'test.csv')
        assert rows[0]['voornaam'] == 'John'
        assert rows[0]['naam'] == 'Smith John'

    def test_prenom_nom_french(self):
        headers = ['date', 'prénom', 'nom', 'email']
        data = _make_csv_rows(headers,
                              [['2026-05-22', 'Marie', 'Dupont', 'marie@example.com']])
        rows, _ = _process_rows(headers, data, 'test.csv')
        assert rows[0]['voornaam'] == 'Marie'
        assert rows[0]['naam'] == 'Dupont Marie'


# ── 3. MyOrganizer-stijl: volledige naam in één kolom ────────────────────────

class TestMyOrganizerFullName:
    HEADERS = ['naam', 'email', 'datum']

    def test_naam_achternaam_voornaam_formaat(self):
        data = _make_csv_rows(self.HEADERS,
                              [['Destuyver Ellen', 'ellen@example.com', '2026-03-15']])
        rows, _ = _process_rows(self.HEADERS, data, 'agenda.csv')
        assert rows[0]['naam'] == 'Destuyver Ellen'
        assert rows[0]['voornaam'] == 'Ellen'
        assert rows[0]['achternaam'] == 'Destuyver'

    def test_email_via_celscan_zonder_emailkolom(self):
        # Geen expliciete e-mailkolomnaam, waarde wordt via scan gevonden
        headers = ['naam', 'contactinfo', 'datum']
        data = _make_csv_rows(headers,
                              [['Pieters Jan', 'jan.pieters@test.be', '01/03/2026']])
        rows, _ = _process_rows(headers, data, 'test.csv')
        assert rows[0]['email'] == 'jan.pieters@test.be'
        assert rows[0]['naam'] == 'Pieters Jan'

    def test_enkel_voornaam_in_naamkolom(self):
        data = _make_csv_rows(self.HEADERS,
                              [['Ellen', 'ellen@example.com', '2026-03-15']])
        rows, _ = _process_rows(self.HEADERS, data, 'agenda.csv')
        assert rows[0]['naam'] == 'Ellen'
        assert rows[0]['voornaam'] == 'Ellen'

    def test_meerdere_namen_ok(self):
        data = _make_csv_rows(self.HEADERS, [
            ['Destuyver Ellen', 'ellen@example.com', '2026-03-15'],
            ['Van Loon Mark',   'mark@example.com',  '2026-03-16'],
        ])
        rows, _ = _process_rows(self.HEADERS, data, 'agenda.csv')
        assert rows[0]['voornaam'] == 'Ellen'
        assert rows[1]['voornaam'] == 'Mark'
        assert rows[1]['naam'] == 'Van Loon Mark'


# ── 4. Datumformaten ──────────────────────────────────────────────────────────

class TestDatumFormaten:
    def test_dd_mm_yyyy_slash(self):
        headers = ['DATUM', 'VOORNAAM', 'ACHTERNAAM', 'E-MAIL']
        data = _make_csv_rows(headers, [['28/05/2026', 'Jan', 'De Vos', 'jan@x.com']])
        rows, _ = _process_rows(headers, data, 'test.csv')
        assert rows[0]['datum_consult'] == '2026-05-28'

    def test_yyyy_mm_dd(self):
        headers = ['datum', 'naam', 'email']
        data = _make_csv_rows(headers, [['2026-05-28', 'Pieters Jan', 'jan@x.com']])
        rows, _ = _process_rows(headers, data, 'test.csv')
        assert rows[0]['datum_consult'] == '2026-05-28'

    def test_dd_mm_yyyy_punt(self):
        headers = ['datum', 'naam', 'email']
        data = _make_csv_rows(headers, [['28.05.2026', 'Pieters Jan', 'jan@x.com']])
        rows, _ = _process_rows(headers, data, 'test.csv')
        assert rows[0]['datum_consult'] == '2026-05-28'


# ── 5. Statistieken in output ─────────────────────────────────────────────────

class TestStats:
    def test_stats_bevat_sheet_key(self):
        headers = ['VOORNAAM', 'ACHTERNAAM', 'E-MAIL']
        data = _make_csv_rows(headers, [['Jan', 'De Vos', 'jan@x.com']])
        _, stats = _process_rows(headers, data, 'test.csv', sheet_name='Resultaten')
        assert stats['sheet'] == 'Resultaten'

    def test_stats_telt_rijen_correct(self):
        headers = ['VOORNAAM', 'ACHTERNAAM', 'E-MAIL']
        data = _make_csv_rows(headers, [
            ['Jan',   'De Vos',    'jan@x.com'],
            ['',      '',          ''],           # leeg
            ['Marie', 'Dupont',    ''],            # geen email
            ['Anna',  'Janssens',  'anna@x.com'],
        ])
        _, stats = _process_rows(headers, data, 'test.csv')
        assert stats['rijen_gelezen'] == 4
        assert stats['rijen_ok'] == 2
        assert stats['rijen_leeg'] == 1
        assert stats['rijen_geen_email'] == 1


# ── 6. Multi-sheet Excel — kiest sheet met meeste e-mails ────────────────────

class TestMultiSheetExcel:
    def test_kiest_resultaten_boven_samenvatting(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()

        # Sheet 1: Samenvatting — geen e-mails
        ws1 = wb.active
        ws1.title = 'Samenvatting'
        ws1.append(['Periode', 'Waarde'])
        ws1.append(['01/05/2026 - 31/05/2026', '221'])

        # Sheet 2: Resultaten — met e-mails
        ws2 = wb.create_sheet('Resultaten')
        ws2.append(['DATUM', 'VOORNAAM', 'ACHTERNAAM', 'E-MAIL'])
        ws2.append(['01/05/2026', 'Lisa', 'Willems', 'lisa@example.com'])
        ws2.append(['02/05/2026', 'Jan',  'Pieters', 'jan@example.com'])

        path = tmp_path / 'test_multi.xlsx'
        wb.save(path)

        rows, stats = load_csv_file(path)
        assert stats['sheet'] == 'Resultaten'
        assert stats['rijen_ok'] == 2
        assert rows[0]['voornaam'] == 'Lisa'
        assert rows[0]['naam'] == 'Willems Lisa'

    def test_sheet_naam_in_stats(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Data'
        ws.append(['naam', 'email'])
        ws.append(['Pieters Jan', 'jan@example.com'])
        path = tmp_path / 'single_sheet.xlsx'
        wb.save(path)

        _, stats = load_csv_file(path)
        assert stats['sheet'] == 'Data'
