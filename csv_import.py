import csv
import io
import re
from pathlib import Path
from datetime import datetime

ENCODINGS_TO_TRY = ['utf-8-sig', 'utf-8', 'cp1252', 'latin-1']
EMAIL_REGEX = re.compile(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$')

NAME_COL_KEYWORDS      = ['naam', 'name', 'patient', 'klant', 'client',
                           'persoon', 'contact', 'omschrijving']
VOORNAAM_COL_KEYWORDS  = ['voornaam', 'first name', 'firstname', 'given name',
                           'roepnaam', 'prénom', 'prenom']
ACHTERNAAM_COL_KEYWORDS = ['achternaam', 'familienaam', 'last name', 'lastname',
                            'surname', 'nom']
EMAIL_COL_KEYWORDS     = ['mail', 'email', 'e-mail', 'emailadres', 'e_mail']
DATE_COL_KEYWORDS      = ['datum', 'date', 'geboortedatum', 'birthday', 'dob']


def try_parse_date(date_str):
    if not date_str:
        return None
    date_str = str(date_str).strip()
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(date_str, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def normalize_email(email):
    return str(email).strip().lower() if email else ''


def extract_voornaam(naam):
    if not naam:
        return ''
    parts = naam.strip().split()
    return parts[-1] if parts else ''


def extract_achternaam(naam):
    if not naam:
        return ''
    parts = naam.strip().split()
    return ' '.join(parts[:-1]) if len(parts) > 1 else naam.strip()


def is_valid_email(email):
    return bool(EMAIL_REGEX.match(str(email).strip()))


def _read_file(path):
    for encoding in ENCODINGS_TO_TRY:
        try:
            with open(path, encoding=encoding, newline='') as f:
                content = f.read()
            return content, encoding
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise ValueError(f"Kan bestand niet decoderen: {path}")


def _detect_delimiter(content):
    """Pick ; or , based on which produces more columns in the header row."""
    first_line = content.split('\n')[0]
    return ';' if first_line.count(';') >= first_line.count(',') else ','


def _xlsx_to_rows(filepath):
    """Return (list-of-lists, sheet_name) from the sheet with the most email-bearing rows."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    best_sheet, best_count, best_rows = None, -1, []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell).strip() if cell is not None else '' for cell in row])
        # Count rows (excluding header) that contain at least one valid email
        email_rows = sum(
            1 for r in rows[1:]
            if any(EMAIL_REGEX.match(str(c).strip().lower()) for c in r if c)
        )
        if email_rows > best_count:
            best_count, best_sheet, best_rows = email_rows, name, rows

    wb.close()
    return best_rows, best_sheet


def _find_col(fieldnames, keywords):
    """Return first column name whose lowercased form contains any keyword."""
    for col in fieldnames:
        col_l = col.strip().lower()
        if any(kw in col_l for kw in keywords):
            return col
    return None


def _extract_email_from_row(values):
    """Scan all cell values and return first valid email found."""
    for v in values:
        v = normalize_email(v)
        if v and is_valid_email(v):
            return v
    return ''


def _extract_name_from_row(row_dict, name_col):
    """Return name: prefer the detected name column, else first longish text cell."""
    if name_col:
        val = row_dict.get(name_col, '').strip()
        if val:
            return val
    # Fallback: first cell with 2+ words that isn't an email or pure number
    for v in row_dict.values():
        v = v.strip()
        if (v and len(v.split()) >= 2
                and not is_valid_email(v)
                and not re.match(r'^[\d./:,\- ]+$', v)):
            return v
    # Last resort: any non-empty text cell
    for v in row_dict.values():
        v = v.strip()
        if v and not is_valid_email(v) and not re.match(r'^[\d./:,\- ]+$', v):
            return v
    return ''


def _process_rows(fieldnames, data_rows, filename, sheet_name=None):
    """
    Format-agnostic processor: scan every cell for emails, find names
    from known column names or heuristics.

    Priority for naam/voornaam/achternaam:
    1. Explicit VOORNAAM + ACHTERNAAM columns → combine as "Achternaam Voornaam"
    2. Single full-name column → keep as-is, split with extract_voornaam/achternaam
    3. Heuristic: first cell with 2+ words that isn't an email or number
    """
    voornaam_col   = _find_col(fieldnames, VOORNAAM_COL_KEYWORDS)
    # Exclude the voornaam column so 'nom' in 'prénom' can't steal the achternaam slot
    _remaining     = [f for f in fieldnames if f != voornaam_col]
    achternaam_col = _find_col(_remaining, ACHTERNAAM_COL_KEYWORDS)
    # Only use generic name_col when split columns are absent
    name_col  = _find_col(fieldnames, NAME_COL_KEYWORDS) if not (voornaam_col and achternaam_col) else None
    date_col  = _find_col(fieldnames, DATE_COL_KEYWORDS)
    datum_col = next((c for c in fieldnames
                      if 'datum' in c.lower() and 'geboorte' not in c.lower()), None)

    stats = {
        'bestand':    filename,
        'sheet':      sheet_name or '',
        'encoding':   'csv',
        'rijen_gelezen':      0,
        'rijen_leeg':         0,
        'rijen_geen_email':   0,
        'rijen_ok':           0,
        'overgeslagen':       [],
    }

    rows = []
    for raw_row in data_rows:
        stats['rijen_gelezen'] += 1
        row = {k.strip(): (v.strip() if v else '') for k, v in raw_row.items() if k}
        row_lower = {k.lower(): v for k, v in row.items()}

        all_values = list(row.values())

        # Skip completely empty rows
        if not any(all_values):
            stats['rijen_leeg'] += 1
            continue

        email = _extract_email_from_row(all_values)
        if not email:
            stats['rijen_geen_email'] += 1
            continue

        # ── Naam resolution ──────────────────────────────────────────────────
        if voornaam_col and achternaam_col:
            # Crossuite-style: separate columns
            vn = row.get(voornaam_col, '').strip()
            an = row.get(achternaam_col, '').strip()
            voornaam  = vn
            achternaam = an
            naam = f'{an} {vn}'.strip() if an else vn
        else:
            # MyOrganizer-style: single name field or heuristic
            naam = _extract_name_from_row(row, name_col)
            if not naam:
                naam = email
            voornaam  = extract_voornaam(naam)
            achternaam = extract_achternaam(naam)
        # ─────────────────────────────────────────────────────────────────────

        geboortedatum = try_parse_date(row_lower.get(date_col.lower(), '') if date_col else '')
        datum_consult = try_parse_date(row_lower.get(datum_col.lower(), '') if datum_col else '')

        stats['rijen_ok'] += 1
        rows.append({
            'naam':          naam,
            'voornaam':      voornaam,
            'achternaam':    achternaam,
            'email':         email,
            'geboortedatum': geboortedatum,
            'datum_consult': datum_consult,
            'telefoon':      row_lower.get('telefoon', ''),
            'gsm':           row_lower.get('gsm nummer', row_lower.get('gsm', '')),
            'agenda':        row_lower.get('agenda', ''),
            'afspraak_type': row_lower.get('afspraak type', row_lower.get('type', '')),
            'bestand':       filename,
        })

    return rows, stats


def _process_csv_content(content, filename):
    delim = _detect_delimiter(content)
    reader = csv.DictReader(io.StringIO(content), delimiter=delim)

    if not reader.fieldnames:
        raise ValueError(f"Geen kolomnamen gevonden in {filename}")

    return _process_rows(reader.fieldnames, list(reader), filename)


def _process_xlsx(filepath):
    all_rows, sheet_name = _xlsx_to_rows(filepath)
    if not all_rows:
        raise ValueError(f"Leeg bestand: {filepath.name}")
    header = all_rows[0]
    data = [dict(zip(header, row)) for row in all_rows[1:]]
    return _process_rows(header, data, filepath.name, sheet_name=sheet_name)


def load_csv_file(filepath):
    path = Path(filepath)
    if path.suffix.lower() in ('.xlsx', '.xls'):
        rows, stats = _process_xlsx(path)
        stats['encoding'] = 'xlsx'
    else:
        content, encoding = _read_file(path)
        rows, stats = _process_csv_content(content, path.name)
        stats['encoding'] = encoding
    return rows, stats


def load_all_csv(input_dir):
    input_dir = Path(input_dir)
    files = sorted([
        f for f in input_dir.iterdir()
        if f.suffix.lower() in ('.csv', '.xlsx', '.xls')
    ])

    if not files:
        raise FileNotFoundError(
            f"Geen bestanden gevonden in '{input_dir.resolve()}'.\n"
            "Upload eerst een CSV of Excel export."
        )

    all_rows, all_stats = [], []
    for f in files:
        rows, stats = load_csv_file(f)
        all_rows.extend(rows)
        all_stats.append(stats)

    return all_rows, all_stats
